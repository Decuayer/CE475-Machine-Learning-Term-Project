import openpyxl
import random
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error
from sklearn.linear_model import Ridge, Lasso
from sklearn.svm import SVR
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.model_selection import cross_val_score


models = {
    "Linear Regression": Ridge(),
    "Lasso Regression": Lasso(),
    "SVR (RBF Kernel)": SVR(kernel='rbf'),
    "Random Forest": RandomForestRegressor(n_estimators=100, random_state=42),
    "Gradient Boosting": GradientBoostingRegressor(random_state=42)
}

class StandartScaler:
    def __init__(self):
        self.mean = None
        self.std = None

    def fit(self, X):
        self.mean = np.mean(X, axis=0)
        self.std = np.std(X, axis=0)
    
    def transform(self, X):
        return (X - self.mean) / self.std
    
    def fit_transform(self, X):
        self.fit(X)
        return self.transform(X)

def load_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Data"]
    
    X_train = np.array([[c.value for c in r] for r in sheet.iter_rows(2, 101, 2, 7)])
    X_test = np.array([[c.value for c in r] for r in sheet.iter_rows(102, 121, 2, 7)])
    y_train = np.array([sheet.cell(r, 8).value for r in range(2, 102)])
        
    return np.array(X_train), np.array(y_train), np.array(X_test)

def plot_scaled_vs(X_train, X_scaled, y_train):
    fig, axes = plt.subplots(1, 2, figsize=(10, 4))

    axes[0].plot(X_scaled, y_train, "*")
    axes[0].set_title("Scaled")

    axes[1].plot(X_train, y_train, "*")
    axes[1].set_title("Not Scaled")
    
    plt.show()

def linear_regression_gd(X, y, lr=0.01, epochs=1000):
    n_samples, n_features = X.shape
    weights = np.zeros((n_features, 1))
    bias = 0
    y = y.reshape(-1, 1)

    for i in range(epochs):
        y_pred = np.dot(X, weights) + bias

        dw = (1 / n_samples) * np.dot(X.T, (y_pred - y))
        db = (1 / n_samples) * np.sum(y_pred - y)

        weights -= lr * dw
        bias -= lr * db

    return weights, bias

def linear_regression_ne(X, y):
    X_b = np.c_[np.ones((X.shape[0], 1)), X]
    
    theta_best = np.linalg.pinv(X_b.T.dot(X_b)).dot(X_b.T).dot(y)

    bias = theta_best[0]
    coefficients = theta_best[1:]

    return coefficients, bias

def compare_models(X_train, y_train, X_test, y_test):
    results = []

def plot_regression_result(X, y, coef, bias):
    fig, axes = plt.subplots(2, 3, figsize=(10, 6))
    axes = axes.flatten()

    for i in range(6):
        axes[i].scatter(X[:, i], y, color='blue', alpha=0.3, label='RealData')
        
        x_range = np.linspace(X[:, i].min(), X[:, i].max(), 100)
        y_line = x_range * coef[i] + bias
        
        axes[i].plot(x_range, y_line, color='red', lw=2, label='Reggression Line')
        axes[i].set_title(f'Feature X-{i+1} vs Y')
        axes[i].legend()

    plt.tight_layout()
    plt.show()

def prepare_data(X, y, test_size=0.2):
    X, y = np.array(X), np.array(y)
    indices = np.arange(X.shape[0])
    np.random.shuffle(indices)

    split_idx = int(len(indices) * (1 - test_size))
    train_idx, test_idx = indices[:split_idx], indices[split_idx:]

    return X[train_idx], X[test_idx], y[train_idx], y[test_idx]

def train_and_evaluate(X_train, y_train, X_test, y_test):
    model = RandomForestRegressor(n_estimators=100, max_depth=5, random_state=42)
    model.fit(X_train, y_train)

    predictions = model.predict(X_test)
    mae = mean_absolute_error(y_test, predictions)

    return model, mae

def prepare_data_sequential(X, y, test_size=0.2):
    X, y = np.array(X), np.array(y)
    split_idx = int(len(X) * (1 - test_size))

    return X[:split_idx], X[split_idx:], y[:split_idx], y[split_idx:]

def plot_RFR_result(y_test, predictions):

    y_test = np.array(y_test).flatten()
    predictions = np.array(predictions).flatten()
    
    if len(y_test) != len(predictions):
        print(f"Hata: y_test boyutu {len(y_test)}, tahmin boyutu {len(predictions)}! Bunlar eşit olmalı.")
        return

    plt.figure(figsize=(10, 6))
    
    plt.scatter(y_test, predictions, color='blue', alpha=0.5, label='Predictions')
    
    ideal = [min(y_test), max(y_test)]
    plt.plot(ideal, ideal, color='red', linestyle='--', label='Perfect Prediction')
    
    plt.xlabel('Actual Values')
    plt.ylabel('Predicted Values')
    plt.title('Random Forest Regressor: Actual vs Predicted')
    plt.legend()
    plt.grid(True)
    plt.show()

def plot_error_dist(y_test, predictions):
    errors = predictions - y_test
    plt.hist(errors, bins=15, color='green', edgecolor='black')
    plt.title('Error Distribution')
    plt.xlabel('Error Amount')
    plt.show()

def compare_models(X_train, y_train, X_test, y_test):
    results = []

    for name, model in models.items():
        model.fit(X_train, y_train)
        preds = model.predict(X_test)
        mae = mean_absolute_error(y_test, preds)
        r2 = model.score(X_test, y_test)
        results.append({"Model": name, "MAE":mae, "R2 Score": r2})
    
    return pd.DataFrame(results)

def evaluate_with_cv(X, y):
    for name, model in models.items():
        scores = cross_val_score(model, X, y, cv=5, scoring='neg_mean_absolute_error')
        print(f"{name} CV MAE: {-scores.mean():.2f}")


if __name__ == "__main__":
    file_path = "data.xlsx" 
    X_train_raw, y_train_raw, X_test_raw = load_data_from_excel(file_path)

    X_scaled = StandartScaler().fit_transform(X_train_raw)

    #X_train, X_test, y_train, y_test = prepare_data(X_scaled, y_train_raw)
    X_train, X_test, y_train, y_test = prepare_data_sequential(X_scaled, y_train_raw)

    #df_scaled = pd.DataFrame(X_scaled, columns=[f'Feature_{i+1}' for i in range(6)])
    #plot_scaled_vs(X_train_raw, X_scaled, y_train_raw)

    coef, bias = linear_regression_ne(X_train, y_train)
    
    plot_regression_result(X_train, y_train, coef, bias)

    


    print("\n--- Model Comparison Results (MAE) ---")
    comparison_df = compare_models(X_train, y_train, X_test, y_test)
    print(comparison_df)

    print("\n--- Model Comparison Results (CV) ---")
    evaluate_with_cv(X_scaled, y_train_raw)

    model, error = train_and_evaluate(X_train, y_train, X_test, y_test)


    predictions = model.predict(X_test)
    plot_RFR_result(y_test, predictions)
    plot_error_dist(y_test, predictions)




