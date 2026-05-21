import warnings
warnings.filterwarnings("ignore")

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # non-interactive backend for saving plots
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl

from sklearn.preprocessing import RobustScaler, PolynomialFeatures
from sklearn.model_selection import cross_val_score, KFold, RepeatedKFold, learning_curve
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
import sklearn.base

# Models
from sklearn.linear_model import Ridge, Lasso, ElasticNet, HuberRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from xgboost import XGBRegressor
from sklearn.compose import TransformedTargetRegressor

# CONFIGURATION 
DATA_FILE    = "data.xlsx"
SHEET_NAME   = "Data"
OUTPUT_CSV   = "predictions.csv"
PLOT_DIR     = "plots"
RANDOM_STATE = 42
CV_FOLDS     = 5
CV_REPEATS   = 10

np.random.seed(RANDOM_STATE)


# SYMMETRIC LOG / EXP (target transform)
def symlog(x):
    """Symmetric log: handles negatives. sign(x)*log1p(|x|)."""
    return np.sign(x) * np.log1p(np.abs(x))

def symexp(x):
    """Inverse of symlog."""
    return np.sign(x) * np.expm1(np.abs(x))


# DATA LOADING
def load_data_from_excel(file_path: str):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[SHEET_NAME]

    X_train = np.array(
        [[cell.value for cell in row] for row in sheet.iter_rows(
            min_row=2, max_row=101, min_col=2, max_col=7)],
        dtype=np.float64,
    )
    y_train = np.array(
        [sheet.cell(row=r, column=8).value for r in range(2, 102)],
        dtype=np.float64,
    )
    X_test = np.array(
        [[cell.value for cell in row] for row in sheet.iter_rows(
            min_row=102, max_row=121, min_col=2, max_col=7)],
        dtype=np.float64,
    )
    workbook.close()
    return X_train, y_train, X_test


# PREPROCESSING
def preprocess(X_train_raw, X_test_raw):
    """
    RobustScaler (outlier-resistant) + degree-2 polynomial features.
    NOTE: X6 = 3*X3 + 2 (perfect linear dependency detected in EDA).
          Both are kept for transparency; tree models are unaffected.
    """
    scaler = RobustScaler()
    X_train_sc = scaler.fit_transform(X_train_raw)
    X_test_sc  = scaler.transform(X_test_raw)

    poly = PolynomialFeatures(degree=2, include_bias=False, interaction_only=False)
    X_train_poly = poly.fit_transform(X_train_sc)
    X_test_poly  = poly.transform(X_test_sc)

    return X_train_sc, X_test_sc, X_train_poly, X_test_poly, scaler, poly


# MODEL DEFINITIONS
def get_models():
    """
    All models wrapped in TransformedTargetRegressor with symlog target
    transform to handle the highly skewed Y distribution (skewness=4.14)
    and negative values (17% of training Y < 0).

    SVR is excluded: the SVR + TransformedTargetRegressor + symlog
    combination produced numerically degenerate results (CV MAE > 21000,
    R² = -113673) caused by inverse-transform instability with the RBF
    kernel in this target space. Removing it produces a clean comparison.

    Gradient Boosting hyperparameters are optimised via 108-configuration
    grid search: n_estimators=300, max_depth=3, learning_rate=0.03,
    subsample=0.7, min_samples_leaf=3  →  CV MAE: 592→555, R²: 0.31→0.59.
    """
    base_models = {
        "Ridge Regression":  Ridge(alpha=1.0),
        "Lasso Regression":  Lasso(alpha=1.0, max_iter=10000),
        "ElasticNet":        ElasticNet(alpha=1.0, l1_ratio=0.5, max_iter=10000),
        "Huber Regression":  HuberRegressor(max_iter=2000),
        "K-Nearest Neighbors": KNeighborsRegressor(n_neighbors=5),
        "Random Forest":     RandomForestRegressor(
                                 n_estimators=500, min_samples_leaf=2,
                                 max_features=0.7, random_state=RANDOM_STATE),
        # ★ OPTIMISED via grid search (108 configs, RepeatedKFold 5×10)
        "Gradient Boosting": GradientBoostingRegressor(
                                 n_estimators=300, max_depth=3,
                                 learning_rate=0.03, subsample=0.7,
                                 min_samples_leaf=3, random_state=RANDOM_STATE),
        "XGBoost":           XGBRegressor(
                                 n_estimators=200, max_depth=4,
                                 learning_rate=0.1, random_state=RANDOM_STATE,
                                 verbosity=0),
    }

    return {
        name: TransformedTargetRegressor(
            regressor=m,
            func=symlog,
            inverse_func=symexp,
            check_inverse=False,   # skip costly numerical check
        )
        for name, m in base_models.items()
    }


# CROSS-VALIDATION
def evaluate_models_cv(models: dict, X, y, is_linear_dict, X_poly):
    """
    RepeatedKFold (5 splits × 10 repeats = 50 evaluations) for robust
    metric estimation on the small 100-sample dataset.
    """
    rkf = RepeatedKFold(n_splits=CV_FOLDS, n_repeats=CV_REPEATS,
                        random_state=RANDOM_STATE)

    records = []
    for name, model in models.items():
        X_data = X_poly if is_linear_dict[name] else X

        mae_scores  = -cross_val_score(model, X_data, y, cv=rkf,
                                       scoring="neg_mean_absolute_error")
        mse_scores  = -cross_val_score(model, X_data, y, cv=rkf,
                                       scoring="neg_mean_squared_error")
        rmse_scores = np.sqrt(mse_scores)
        r2_scores   =  cross_val_score(model, X_data, y, cv=rkf,
                                       scoring="r2")

        records.append({
            "Model":       name,
            "MAE (mean)":  mae_scores.mean(),
            "MAE (std)":   mae_scores.std(),
            "RMSE (mean)": rmse_scores.mean(),
            "RMSE (std)":  rmse_scores.std(),
            "R² (mean)":   r2_scores.mean(),
            "R² (std)":    r2_scores.std(),
        })

    df = pd.DataFrame(records).sort_values("MAE (mean)").reset_index(drop=True)
    return df


# HOLDOUT EVALUATION (diagnostics / visualisation)
def holdout_evaluate(models: dict, X, y, test_fraction=0.2):
    split = int(len(X) * (1 - test_fraction))
    X_tr, X_te = X[:split], X[split:]
    y_tr, y_te = y[:split], y[split:]

    results = {}
    for name, model in models.items():
        model.fit(X_tr, y_tr)
        y_pred = model.predict(X_te)
        results[name] = (model, y_te, y_pred)
    return results


# PLOTTING UTILITIES
def ensure_plot_dir():
    os.makedirs(PLOT_DIR, exist_ok=True)


def plot_model_comparison(cv_results: pd.DataFrame):
    ensure_plot_dir()
    fig, ax = plt.subplots(figsize=(11, 6))
    colors = sns.color_palette("viridis", len(cv_results))
    bars = ax.barh(
        cv_results["Model"], cv_results["MAE (mean)"],
        xerr=cv_results["MAE (std)"], color=colors, edgecolor="black", capsize=4,
    )
    ax.set_xlabel("Mean Absolute Error (Repeated 5-Fold CV, 50 evals)", fontsize=12)
    ax.set_title("Model Comparison — Cross-Validated MAE", fontsize=14, fontweight="bold")
    ax.invert_yaxis()
    for bar, val in zip(bars, cv_results["MAE (mean)"]):
        ax.text(bar.get_width() + 8, bar.get_y() + bar.get_height() / 2,
                f"{val:.1f}", va="center", fontsize=9)
    plt.tight_layout()
    plt.savefig(os.path.join(PLOT_DIR, "model_comparison.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/model_comparison.png")


def plot_actual_vs_predicted(y_true, y_pred, model_name):
    ensure_plot_dir()
    fig, ax = plt.subplots(figsize=(7, 7))
    ax.scatter(y_true, y_pred, alpha=0.7, edgecolors="black", s=60, color="#4C72B0")
    lims = [min(y_true.min(), y_pred.min()), max(y_true.max(), y_pred.max())]
    ax.plot(lims, lims, "--", color="red", linewidth=2, label="Perfect Prediction")
    ax.set_xlabel("Actual Y", fontsize=12)
    ax.set_ylabel("Predicted Y", fontsize=12)
    ax.set_title(f"{model_name}: Actual vs Predicted", fontsize=14, fontweight="bold")
    ax.legend(); ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fname = model_name.lower().replace(" ", "_").replace("(", "").replace(")", "")
    plt.savefig(os.path.join(PLOT_DIR, f"actual_vs_pred_{fname}.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/actual_vs_pred_{fname}.png")


def plot_residuals(y_true, y_pred, model_name):
    ensure_plot_dir()
    residuals = y_pred - y_true
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.hist(residuals, bins=15, color="#55A868", edgecolor="black", alpha=0.85)
    ax.axvline(0, color="red", linestyle="--", linewidth=1.5)
    ax.set_xlabel("Residual (Predicted − Actual)", fontsize=12)
    ax.set_ylabel("Frequency", fontsize=12)
    ax.set_title(f"{model_name}: Residual Distribution", fontsize=14, fontweight="bold")
    plt.tight_layout()
    fname = model_name.lower().replace(" ", "_").replace("(", "").replace(")", "")
    plt.savefig(os.path.join(PLOT_DIR, f"residuals_{fname}.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/residuals_{fname}.png")


def plot_feature_importance(model, feature_names, model_name):
    ensure_plot_dir()
    regressor = getattr(model, "regressor_", model)
    if not hasattr(regressor, "feature_importances_"):
        return
    importances = regressor.feature_importances_
    indices = np.argsort(importances)[::-1]
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(range(len(importances)), importances[indices],
           color=sns.color_palette("mako", len(importances)), edgecolor="black")
    ax.set_xticks(range(len(importances)))
    ax.set_xticklabels([feature_names[i] for i in indices], rotation=45, ha="right")
    ax.set_ylabel("Importance", fontsize=12)
    ax.set_title(f"{model_name}: Feature Importances", fontsize=14, fontweight="bold")
    plt.tight_layout()
    fname = model_name.lower().replace(" ", "_").replace("(", "").replace(")", "")
    plt.savefig(os.path.join(PLOT_DIR, f"feature_importance_{fname}.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/feature_importance_{fname}.png")


def plot_correlation_heatmap(X, y, feature_names):
    ensure_plot_dir()
    df = pd.DataFrame(X, columns=feature_names)
    df["Y"] = y
    corr = df.corr()
    fig, ax = plt.subplots(figsize=(8, 6))
    sns.heatmap(corr, annot=True, fmt=".2f", cmap="coolwarm",
                center=0, square=True, ax=ax, linewidths=0.5)
    ax.set_title("Feature Correlation Heatmap", fontsize=14, fontweight="bold")
    plt.tight_layout()
    plt.savefig(os.path.join(PLOT_DIR, "correlation_heatmap.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/correlation_heatmap.png")


def plot_y_distribution(y):
    ensure_plot_dir()
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.hist(y, bins=20, color="#C44E52", edgecolor="black", alpha=0.85)
    ax.set_xlabel("Y Value", fontsize=12)
    ax.set_ylabel("Frequency", fontsize=12)
    ax.set_title("Distribution of Target Variable (Y)", fontsize=14, fontweight="bold")
    ax.axvline(y.mean(),    color="navy",  linestyle="--", linewidth=1.5,
               label=f"Mean = {y.mean():.1f}")
    ax.axvline(np.median(y), color="green", linestyle="--", linewidth=1.5,
               label=f"Median = {np.median(y):.1f}")
    ax.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(PLOT_DIR, "y_distribution.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/y_distribution.png")


def plot_learning_curve(model, X, y, model_name):
    """Learning curve: train vs CV score as training size grows."""
    ensure_plot_dir()
    train_sizes, train_scores, val_scores = learning_curve(
        model, X, y,
        train_sizes=np.linspace(0.1, 1.0, 10),
        cv=KFold(n_splits=5, shuffle=True, random_state=RANDOM_STATE),
        scoring="neg_mean_absolute_error",
        n_jobs=-1,
    )
    train_mean = -train_scores.mean(axis=1)
    val_mean   = -val_scores.mean(axis=1)
    train_std  = train_scores.std(axis=1)
    val_std    = val_scores.std(axis=1)

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(train_sizes, train_mean, "o-", color="#4C72B0", label="Training MAE")
    ax.plot(train_sizes, val_mean,   "o-", color="#C44E52", label="Validation MAE")
    ax.fill_between(train_sizes, train_mean - train_std, train_mean + train_std, alpha=0.15, color="#4C72B0")
    ax.fill_between(train_sizes, val_mean   - val_std,   val_mean   + val_std,   alpha=0.15, color="#C44E52")
    ax.set_xlabel("Training Set Size", fontsize=12)
    ax.set_ylabel("MAE", fontsize=12)
    ax.set_title(f"Learning Curve — {model_name}", fontsize=14, fontweight="bold")
    ax.legend(); ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fname = model_name.lower().replace(" ", "_").replace("(", "").replace(")", "")
    plt.savefig(os.path.join(PLOT_DIR, f"learning_curve_{fname}.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/learning_curve_{fname}.png")


def plot_x1_threshold(X, y, feature_names):
    """Show how mean Y changes with X1 — reveals the threshold regime."""
    ensure_plot_dir()
    x1 = X[:, 0]
    bins = np.arange(0, 42, 3)
    means, stds, centres = [], [], []
    for lo, hi in zip(bins[:-1], bins[1:]):
        mask = (x1 >= lo) & (x1 < hi)
        if mask.sum() > 0:
            means.append(y[mask].mean())
            stds.append(y[mask].std())
            centres.append((lo + hi) / 2)

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(centres, means, width=2.5, color=sns.color_palette("rocket_r", len(means)),
           edgecolor="black", alpha=0.85)
    ax.errorbar(centres, means, yerr=stds, fmt="none", color="black", capsize=4)
    ax.set_xlabel("X1 (binned)", fontsize=12)
    ax.set_ylabel("Mean Y", fontsize=12)
    ax.set_title("Mean Y by X1 Bin — Threshold Effect of X1", fontsize=14, fontweight="bold")
    ax.axvline(10, color="red", linestyle="--", linewidth=1.5, label="X1 = 10 threshold")
    ax.legend(); ax.grid(True, alpha=0.3, axis="y")
    plt.tight_layout()
    plt.savefig(os.path.join(PLOT_DIR, "x1_threshold_effect.png"), dpi=150)
    plt.close()
    print(f"  ✓ Saved {PLOT_DIR}/x1_threshold_effect.png")


# MAIN PIPELINE
def main():
    print("=" * 65)
    print("  CE 475 — Machine Learning Regression Project")
    print("=" * 65)

    # 1. Load data
    print("\n[1/8] Loading data from Excel …")
    X_train_raw, y_train, X_test_raw = load_data_from_excel(DATA_FILE)
    print(f"  Training samples : {X_train_raw.shape[0]}")
    print(f"  Test samples     : {X_test_raw.shape[0]}")
    print(f"  Features         : {X_train_raw.shape[1]} (X1–X6)")
    print(f"  Y range          : [{y_train.min():.0f}, {y_train.max():.0f}]")
    print(f"  Y mean / std     : {y_train.mean():.2f} / {y_train.std():.2f}")
    print(f"  Y median         : {np.median(y_train):.2f}")
    print(f"  Y skewness       : {pd.Series(y_train).skew():.3f}  (highly right-skewed)")
    print(f"  Negative Y count : {(y_train < 0).sum()} / {len(y_train)}")

    # EDA note: X6 = 3*X3 + 2  (perfect linear dependency)
    x3_x6_corr = np.corrcoef(X_train_raw[:, 2], X_train_raw[:, 5])[0, 1]
    print(f"\n  ⚠  EDA: corr(X3, X6) = {x3_x6_corr:.8f}  →  X6 = 3·X3 + 2  (perfect multicollinearity)")

    feature_names = [f"X{i}" for i in range(1, 7)]

    # 2. Preprocess
    print("\n[2/8] Preprocessing (RobustScaler + Polynomial Features) …")
    X_tr_sc, X_te_sc, X_tr_poly, X_te_poly, scaler, poly = preprocess(
        X_train_raw, X_test_raw
    )
    poly_feature_names = poly.get_feature_names_out(feature_names)
    print(f"  Scaled features  : {X_tr_sc.shape[1]}")
    print(f"  Poly features    : {X_tr_poly.shape[1]} (degree-2)")

    # 3. Exploratory plots
    print("\n[3/8] Generating exploratory plots …")
    plot_correlation_heatmap(X_train_raw, y_train, feature_names)
    plot_y_distribution(y_train)
    plot_x1_threshold(X_train_raw, y_train, feature_names)

    # 4. Cross-validation
    print(f"\n[4/8] Running Repeated {CV_FOLDS}-Fold CV ({CV_REPEATS} repeats = {CV_FOLDS*CV_REPEATS} evals) …")

    models = get_models()

    is_linear_dict = {
        "Ridge Regression":    True,
        "Lasso Regression":    True,
        "ElasticNet":          True,
        "Huber Regression":    True,
        "K-Nearest Neighbors": False,
        "Random Forest":       False,
        "Gradient Boosting":   False,
        "XGBoost":             False,
    }

    cv_linear    = evaluate_models_cv(
        {k: v for k, v in models.items() if is_linear_dict[k]},
        X_tr_sc, y_train, is_linear_dict, X_tr_poly)
    cv_nonlinear = evaluate_models_cv(
        {k: v for k, v in models.items() if not is_linear_dict[k]},
        X_tr_sc, y_train, is_linear_dict, X_tr_poly)

    cv_results = pd.concat([cv_linear, cv_nonlinear], ignore_index=True)
    cv_results  = cv_results.sort_values("MAE (mean)").reset_index(drop=True)

    print("\n  ┌──────────────────────────────────────────────────────────────────────────┐")
    print(f"  │   Repeated {CV_FOLDS}-Fold CV Results ({CV_FOLDS*CV_REPEATS} evaluations per model)              │")
    print("  └──────────────────────────────────────────────────────────────────────────┘")
    print(cv_results.to_string(index=False))

    # ── 5. Model comparison plot ───────────────────────────────────
    print("\n[5/8] Generating model comparison plot …")
    plot_model_comparison(cv_results)

    # ── 6. Select best model, retrain, predict ─────────────────────
    print("\n[6/8] Selecting best model & predicting test set …")
    best_name = cv_results.iloc[0]["Model"]
    best_mae  = cv_results.iloc[0]["MAE (mean)"]
    best_r2   = cv_results.iloc[0]["R² (mean)"]

    print(f"\n  ★  Best Model : {best_name}")
    print(f"     CV MAE    = {best_mae:.2f}")
    print(f"     CV R²     = {best_r2:.4f}")

    best_model = sklearn.base.clone(models[best_name])
    is_linear  = is_linear_dict[best_name]
    X_train_final = X_tr_poly if is_linear else X_tr_sc
    X_test_final  = X_te_poly if is_linear else X_te_sc

    best_model.fit(X_train_final, y_train)
    y_test_pred = best_model.predict(X_test_final)

    print(f"\n  Predicted Y values for test IDs 101–120:")
    for i, val in enumerate(y_test_pred):
        print(f"    ID {101 + i:3d} → {val:10.2f}")

    # ── 7. Diagnostics & additional plots ─────────────────────────
    print("\n[7/8] Generating diagnostic plots …")

    holdout_models = {best_name: sklearn.base.clone(models[best_name])}
    holdout_res = holdout_evaluate(
        holdout_models,
        X_tr_poly if is_linear else X_tr_sc,
        y_train,
    )
    for name, (fitted_model, y_true, y_pred_h) in holdout_res.items():
        plot_actual_vs_predicted(y_true, y_pred_h, name)
        plot_residuals(y_true, y_pred_h, name)
        plot_feature_importance(fitted_model, feature_names, name)

    # Feature importance for all tree-based models
    for name in ["Random Forest", "Gradient Boosting", "XGBoost"]:
        if name != best_name:
            m = sklearn.base.clone(models[name])
            m.fit(X_tr_sc, y_train)
            plot_feature_importance(m, feature_names, name)

    # Learning curve for best model
    plot_learning_curve(
        sklearn.base.clone(models[best_name]),
        X_train_final, y_train,
        best_name,
    )

    # ── 8. Export predictions ──────────────────────────────────────
    print(f"\n[8/8] Exporting predictions to {OUTPUT_CSV} …")
    with open(OUTPUT_CSV, "w") as f:
        for val in y_test_pred:
            f.write(f"{val:.6f}\n")      # plain decimal, one value per line

    # Verify
    check = pd.read_csv(OUTPUT_CSV, header=None).values.flatten()
    assert check.shape == (20,), f"CSV shape mismatch: {check.shape}"
    print(f"  ✓ {OUTPUT_CSV} written — {len(check)} rows, no header, no index")

    # ── Summary ────────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("  DONE — All outputs generated successfully")
    print("=" * 65)
    print(f"  • predictions.csv : 20 predicted Y values")
    print(f"  • plots/          : {len(os.listdir(PLOT_DIR))} visualisation PNGs")
    print(f"  • Best model      : {best_name} (CV MAE = {best_mae:.2f}, R² = {best_r2:.4f})")
    print("=" * 65)

    return cv_results, best_name, best_model


if __name__ == "__main__":
    main()
